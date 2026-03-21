"""Shared utilities for financial services verifiers.

All verifiers load a golden xlsx and compare the agent's submitted Google Sheet
against it. This module contains the common logic: helpers, golden loading,
Google Workspace retrieval (with xlsx download), and the generic check runner.
"""

from __future__ import annotations

import base64
import contextlib
import json
import logging
import os
import re
import tempfile
import urllib.parse
from pathlib import Path
from typing import Any

import openpyxl
import requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

from collinear.core.models import Action
from collinear.core.run_artifacts import RunArtifacts
from collinear.core.tool_calling_client import ToolCallingClient
from collinear.core.verifier import VerifierResult

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
# NUMERIC HELPERS
# ═══════════════════════════════════════════════════════════════════════


def try_float(v: object) -> float | None:
    """Parse a value as float, handling common financial formats."""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        cleaned = (
            v.replace(",", "")
            .replace("$", "")
            .replace("(", "-")
            .replace(")", "")
            .replace("%", "")
            .strip()
        )
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    return None


def numbers_close(a: float, b: float, tol_pct: float = 0.10, abs_tol: float = 2) -> bool:
    """Check if two numbers are within tolerance (relative or absolute)."""
    if a is None or b is None:
        return False
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return abs(a - b) <= abs_tol
    return abs(a - b) / max(abs(a), abs(b)) <= tol_pct


def find_number(
    all_numbers: list[float],
    target: float,
    tol_pct: float = 0.10,
    abs_tol: float = 2,
) -> bool:
    """Check if target exists in all_numbers, also at common scales ($M↔$K, %↔decimal)."""
    for n in all_numbers:
        if numbers_close(n, target, tol_pct, abs_tol):
            return True
    scales = [1000, 1 / 1000]
    if 0 < abs(target) < 1:
        scales.extend([100, 1 / 100])
    for scale in scales:
        for n in all_numbers:
            if numbers_close(n, target * scale, tol_pct, abs_tol):
                return True
    return False


def find_value(
    all_numbers: list[float],
    all_text: str,
    target: float,
    tol_pct: float = 0.10,
    abs_tol: float = 2,
) -> bool:
    """Find a numeric value in numbers or in text (for rates shown as %)."""
    if find_number(all_numbers, target, tol_pct, abs_tol):
        return True
    # For rates stored as decimals in golden (e.g. 0.16), also check
    # percentage text in submitted (e.g. "16" or "16%").
    if 0 < abs(target) < 1:
        pct = target * 100
        for fmt in [f"{pct:.4f}", f"{pct:.3f}", f"{pct:.2f}", f"{pct:.1f}"]:
            trimmed = fmt.rstrip("0").rstrip(".")
            if trimmed in all_text:
                return True
    return False


# ═══════════════════════════════════════════════════════════════════════
# TEXT / DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════


def build_all_text(sheets_data: dict[str, list[list[Any]]]) -> str:
    """Concatenate all cell values into one lowercase text blob."""
    parts: list[str] = []
    for rows in sheets_data.values():
        for row in rows:
            parts.extend(str(v).lower() for v in row if v is not None and v != "")
    return " " + " ".join(parts)


def extract_all_numbers(sheets_data: dict[str, list[list[Any]]]) -> list[float]:
    """Extract all numeric values from sheet data."""
    numbers: list[float] = []
    for rows in sheets_data.values():
        for row in rows:
            for v in row:
                num = try_float(v)
                if num is not None:
                    numbers.append(num)
    return numbers


def count_non_empty(sheets_data: dict[str, list[list[Any]]]) -> int:
    """Count non-empty cells across all sheets."""
    count = 0
    for rows in sheets_data.values():
        for row in rows:
            for v in row:
                if v is not None and v != "":
                    count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════
# GOLDEN / XLSX LOADING
# ═══════════════════════════════════════════════════════════════════════


def load_xlsx(
    path: Path,
) -> tuple[list[str], dict[str, list[list[Any]]]]:
    """Load an xlsx file and return (sheet_names, sheets_data).

    Uses openpyxl with data_only=True so formulas are resolved to values.
    """
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = list(wb.sheetnames)
    sheets_data: dict[str, list[list[Any]]] = {}
    for name in sheet_names:
        ws = wb[name]
        sheets_data[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return sheet_names, sheets_data


def get_col_headers(rows: list[list[Any]], scan_rows: int = 5) -> dict[int, str]:
    """Extract column headers from first N rows."""
    headers: dict[int, str] = {}
    for row in rows[:scan_rows]:
        if row is None:
            continue
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip() and i not in headers:
                headers[i] = v.strip()[:40]
    return headers


def get_row_label(row: list[Any]) -> str:
    """First substantial text cell in a row, used as a label."""
    for v in row:
        if isinstance(v, str) and len(v.strip()) > 2:
            return v.strip()[:50]
    return ""


def extract_labeled_numerics(
    sheets_data: dict[str, list[list[Any]]],
) -> list[tuple[str, float]]:
    """Extract (label, value) pairs from an xlsx, deduped by value.

    Labels are auto-generated from sheet name + row context + column headers.
    """
    results: list[tuple[str, float]] = []
    seen: set[float] = set()

    for sheet_name, rows in sheets_data.items():
        if not rows:
            continue
        headers = get_col_headers(rows)
        for row in rows:
            row_label = get_row_label(row)
            for col_idx, v in enumerate(row):
                if not isinstance(v, (int, float)):
                    continue
                val = float(v)
                if val == 0:
                    continue
                if isinstance(v, int) and 1 <= abs(v) <= 9:
                    continue
                key = round(val, 8)
                if key in seen:
                    continue
                seen.add(key)
                col_head = headers.get(col_idx, "")
                parts = [sheet_name]
                if row_label:
                    parts.append(row_label)
                if col_head and col_head != row_label:
                    parts.append(col_head)
                results.append((" / ".join(parts), val))

    return results


# ═══════════════════════════════════════════════════════════════════════
# GOOGLE WORKSPACE — FIND SPREADSHEET
# ═══════════════════════════════════════════════════════════════════════


def get_workspace_spreadsheet(
    client: ToolCallingClient,
) -> tuple[str | None, str]:
    """Find the output spreadsheet in the workspace folder.

    Returns (spreadsheet_id, error_message).
    """
    result = client.step(Action(tool_name="get_workspace_folder", parameters={}))
    if result.observation.is_error:
        return None, f"Failed to get workspace folder: {result.observation.text}"

    workspace_info = result.observation.structured_content
    if not workspace_info or not isinstance(workspace_info, dict):
        return None, f"Invalid workspace folder response: {result.observation.text}"

    folder_id = workspace_info.get("folder_id")
    if not folder_id:
        return None, "Workspace folder ID not found in response"

    result = client.step(Action(tool_name="list_drive_items", parameters={"folder_id": folder_id}))
    if result.observation.is_error:
        return None, f"Failed to list workspace files: {result.observation.text}"

    files = result.observation.structured_content
    if not files:
        return None, "No files found in workspace folder"

    file_list = files if isinstance(files, list) else files.get("files", [])

    for f in file_list:
        name = (f.get("name") or "").lower()
        if "output_file" in name:
            file_id = f.get("id")
            if file_id:
                return file_id, ""

    for f in file_list:
        mime = (f.get("mimeType") or f.get("mime_type") or "").lower()
        if "spreadsheet" in mime:
            file_id = f.get("id")
            if file_id:
                return file_id, ""

    available = [f.get("name", "?") for f in file_list]
    return None, f"output_file not found. Available files: {available}"


# ═══════════════════════════════════════════════════════════════════════
# GOOGLE WORKSPACE — READ SUBMITTED SHEET
# ═══════════════════════════════════════════════════════════════════════


def _download_submitted_xlsx(
    client: ToolCallingClient, spreadsheet_id: str
) -> tuple[list[str], dict[str, list[list[Any]]], str]:
    """Download the submitted Google Sheet as xlsx and read with openpyxl.

    This gives type-consistent data with the golden (both read by openpyxl).
    """
    result = client.step(
        Action(
            tool_name="get_drive_file_download_url",
            parameters={
                "file_id": spreadsheet_id,
                "export_format": "xlsx",
            },
        )
    )
    if result.observation.is_error:
        return [], {}, f"Download URL request failed: {result.observation.text}"

    # Extract URL from response (may be in structured_content or text)
    url = None
    sc = result.observation.structured_content
    if isinstance(sc, dict):
        url = sc.get("url") or sc.get("download_url") or sc.get("downloadUrl")
    if not url:
        text = result.observation.text or str(sc or "")
        urls = re.findall(r"https?://[^\s<>\"']+", text)
        url = urls[0] if urls else None

    if not url:
        return [], {}, f"Could not extract URL from response: {result.observation.text}"

    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ("https", "http"):
        return [], {}, f"URL scheme not permitted: {parsed.scheme}"

    tmp_path = None
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="wb") as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        sheet_names, sheets_data = load_xlsx(Path(tmp_path))
    except Exception as e:
        return [], {}, f"Download/parse failed: {e}"
    else:
        return sheet_names, sheets_data, ""
    finally:
        if tmp_path:
            with contextlib.suppress(OSError):
                Path(tmp_path).unlink()


def _read_via_api(
    client: ToolCallingClient, spreadsheet_id: str
) -> tuple[list[str], dict[str, list[list[Any]]], str]:
    """Read sheet data via the Google Sheets API (fallback).

    Values come back as formatted strings, parsed by try_float().
    """
    result = client.step(
        Action(
            tool_name="get_spreadsheet_info",
            parameters={"spreadsheet_id": spreadsheet_id},
        )
    )
    if result.observation.is_error:
        return [], {}, f"Failed to get spreadsheet info: {result.observation.text}"

    info = result.observation.structured_content
    if not info or not isinstance(info, dict):
        try:
            info = json.loads(result.observation.text)
        except (json.JSONDecodeError, TypeError):
            return [], {}, f"Invalid spreadsheet info: {result.observation.text}"

    sheets = info.get("sheets", [])
    sheet_names: list[str] = []
    for s in sheets:
        if isinstance(s, dict):
            props = s.get("properties", s)
            title = props.get("title") or props.get("name") or ""
            if title:
                sheet_names.append(title)
        elif isinstance(s, str):
            sheet_names.append(s)

    if not sheet_names:
        return [], {}, "No sheets found in spreadsheet"

    sheets_data: dict[str, list[list[Any]]] = {}
    for sheet_name in sheet_names:
        result = client.step(
            Action(
                tool_name="read_sheet_values",
                parameters={
                    "spreadsheet_id": spreadsheet_id,
                    "range": f"'{sheet_name}'",
                },
            )
        )
        if result.observation.is_error:
            logger.warning("Failed to read sheet %s: %s", sheet_name, result.observation.text)
            sheets_data[sheet_name] = []
            continue

        content = result.observation.structured_content
        if isinstance(content, list):
            sheets_data[sheet_name] = content
        elif isinstance(content, dict):
            sheets_data[sheet_name] = content.get("values", [])
        else:
            try:
                parsed = json.loads(result.observation.text)
                if isinstance(parsed, list):
                    sheets_data[sheet_name] = parsed
                elif isinstance(parsed, dict):
                    sheets_data[sheet_name] = parsed.get("values", [])
                else:
                    sheets_data[sheet_name] = []
            except (json.JSONDecodeError, TypeError):
                sheets_data[sheet_name] = []

    return sheet_names, sheets_data, ""


def read_submitted_spreadsheet(
    client: ToolCallingClient, spreadsheet_id: str
) -> tuple[list[str], dict[str, list[list[Any]]], str]:
    """Read the submitted spreadsheet. Tries xlsx download first, falls back to API.

    The xlsx download gives type-consistent data with the golden file (both
    read by openpyxl → native int/float/str). The API fallback returns
    formatted strings which are parsed by try_float().
    """
    sheet_names, sheets_data, error = _download_submitted_xlsx(client, spreadsheet_id)
    if not error and sheet_names:
        logger.info("Successfully read submitted sheet via xlsx download")
        return sheet_names, sheets_data, ""

    logger.warning("xlsx download failed (%s), falling back to Sheets API", error)
    return _read_via_api(client, spreadsheet_id)


# ═══════════════════════════════════════════════════════════════════════
# GENERIC CHECK RUNNER
# ═══════════════════════════════════════════════════════════════════════


def run_checks(
    sheet_names: list[str],
    sheets_data: dict[str, list[list[Any]]],
    golden_sheet_names: list[str],
    golden_sheets_data: dict[str, list[list[Any]]],
    concept_candidates: list[tuple[str, str]],
    num_tol_pct: float = 0.10,
    num_abs_tol: float = 2,
    rate_tol_pct: float = 0.05,
    rate_abs_tol: float = 0.0005,
    min_cells: int = 100,
    text_coverage_threshold: float = 0.3,
    cross_checks: list[Any] | None = None,
) -> list[dict[str, Any]]:
    """Run all checks comparing submitted against golden reference.

    Parameters
    ----------
    sheet_names
        Names of sheets in the submitted spreadsheet.
    sheets_data
        Cell data keyed by sheet name for the submitted spreadsheet.
    golden_sheet_names
        Expected sheet names from the golden reference.
    golden_sheets_data
        Cell data keyed by sheet name for the golden reference.
    concept_candidates
        Domain keywords to check — only active if they appear in golden text.
    num_tol_pct
        Relative tolerance for dollar-scale values.
    num_abs_tol
        Absolute tolerance for dollar-scale values.
    rate_tol_pct
        Relative tolerance for rate/decimal-scale values.
    rate_abs_tol
        Absolute tolerance for rate/decimal-scale values.
    min_cells
        Minimum non-empty cells for the content check.
    text_coverage_threshold
        Fraction of golden text terms that must match.
    cross_checks
        Optional list of additional cross-check functions that accept
        (sheets_data, all_numbers, all_text) and return criteria tuples.

    """
    results: list[dict[str, Any]] = []

    def add(criteria: str, passed: bool) -> None:
        results.append({"criteria": criteria, "pass": bool(passed)})

    # ── 1. STRUCTURE (from golden sheet names) ──
    add("File opens as valid spreadsheet", passed=True)

    submitted_lower = [s.lower() for s in sheet_names]
    for gs in golden_sheet_names:
        gs_words = [w for w in gs.lower().replace("_", " ").replace("&", " ").split() if len(w) > 2]
        found = any(all(w in sn for w in gs_words) for sn in submitted_lower)
        add(f"Sheet present: {gs}", found)

    add(
        f"At least {len(golden_sheet_names)} sheets",
        len(sheet_names) >= len(golden_sheet_names),
    )

    # ── 2. NUMERIC DATA (from golden xlsx) ──
    golden_numerics = extract_labeled_numerics(golden_sheets_data)
    submitted_nums = extract_all_numbers(sheets_data)
    submitted_text = build_all_text(sheets_data)

    for label, val in golden_numerics:
        if 0 < abs(val) < 1:
            found = find_value(submitted_nums, submitted_text, val, rate_tol_pct, rate_abs_tol)
        else:
            found = find_value(submitted_nums, submitted_text, val, num_tol_pct, num_abs_tol)
        add(f"Golden value: {label} ≈ {val:g}", found)

    # ── 3. CONCEPT KEYWORDS (golden-filtered) ──
    golden_text = build_all_text(golden_sheets_data)
    active_concepts = [(kw, desc) for kw, desc in concept_candidates if kw in golden_text]
    for kw, desc in active_concepts:
        add(f"Concept: {desc}", kw in submitted_text)

    # ── 4. TEXT COVERAGE (per golden sheet) ──
    for gs_name, gs_rows in golden_sheets_data.items():
        terms: set[str] = set()
        for row in gs_rows:
            for v in row:
                if isinstance(v, str) and len(v.strip()) > 10:
                    terms.add(v.strip().lower())
        if terms:
            matched = sum(1 for t in terms if t in submitted_text)
            pct = matched / len(terms)
            add(
                f"Text coverage [{gs_name}]: {matched}/{len(terms)} references",
                pct >= text_coverage_threshold,
            )

    # ── 5. FORMATTING ──
    add(
        "Number formats ($ or comma separators)",
        "$" in submitted_text or "," in submitted_text,
    )
    add("Percentage formats present", "%" in submitted_text)
    add(
        f"Substantial content (>={min_cells} non-empty cells)",
        count_non_empty(sheets_data) >= min_cells,
    )

    # ── 6. CROSS-VALIDATION CHECKS ──
    if cross_checks:
        for check_fn in cross_checks:
            for criteria, passed, _detail in check_fn(sheets_data, submitted_nums, submitted_text):
                add(criteria, passed)

    return results


# ═══════════════════════════════════════════════════════════════════════
# DIRECT GOOGLE DRIVE DOWNLOAD (no tool server required)
# ═══════════════════════════════════════════════════════════════════════

_CREATED_SHEET_RE = re.compile(
    r"Successfully created spreadsheet.{0,80}ID:\s*([A-Za-z0-9_-]{25,})",
    re.IGNORECASE,
)
_SHEET_URL_RE = re.compile(r"spreadsheets/d/([A-Za-z0-9_-]{25,})")


def extract_spreadsheet_id_from_messages(messages: list[dict[str, Any]]) -> str | None:
    """Scan agent messages for a created Google Sheets ID.

    Looks for 'Successfully created spreadsheet ... ID: <id>' first,
    then falls back to any spreadsheets/d/<id> URL in a tool response.
    """
    # Primary: creation confirmation message
    for msg in messages:
        content = msg.get("content", "")
        if isinstance(content, str):
            m = _CREATED_SHEET_RE.search(content)
            if m:
                return m.group(1)
    # Fallback: any Sheets URL in tool messages
    for msg in messages:
        if msg.get("role") != "tool":
            continue
        content = msg.get("content", "")
        if isinstance(content, str):
            m = _SHEET_URL_RE.search(content)
            if m:
                return m.group(1)
    return None


def _download_xlsx_direct(
    spreadsheet_id: str,
) -> tuple[list[str], dict[str, list[list[Any]]], str]:
    """Download a Google Sheet as xlsx using credentials from the environment.

    Reads GOOGLE_WORKSPACE_CREDENTIALS_CONFIG (base64-encoded JSON with
    refresh_token, client_id, client_secret) and refreshes the OAuth token
    before downloading.  No tool server needed.
    """
    raw = os.environ.get("GOOGLE_WORKSPACE_CREDENTIALS_CONFIG")
    if not raw:
        return [], {}, "GOOGLE_WORKSPACE_CREDENTIALS_CONFIG not set in environment"

    try:
        creds_data = json.loads(base64.b64decode(raw))
        creds = Credentials(
            token=None,
            refresh_token=creds_data["refresh_token"],
            token_uri=creds_data["token_uri"],
            client_id=creds_data["client_id"],
            client_secret=creds_data["client_secret"],
            scopes=creds_data["scopes"],
        )
        creds.refresh(Request())
    except Exception as e:
        return [], {}, f"Failed to obtain Google credentials: {e}"

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    tmp_path = None
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {creds.token}"}, timeout=30)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="wb") as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        sheet_names, sheets_data = load_xlsx(Path(tmp_path))
    except Exception as e:
        return [], {}, f"Direct download/parse failed: {e}"
    else:
        return sheet_names, sheets_data, ""
    finally:
        if tmp_path:
            with contextlib.suppress(OSError):
                Path(tmp_path).unlink()


# ═══════════════════════════════════════════════════════════════════════
# GENERIC VERIFY ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════


def run_verifier(
    run_artifacts: RunArtifacts,
    golden_path: Path,
    concept_candidates: list[tuple[str, str]],
    num_tol_pct: float = 0.10,
    num_abs_tol: float = 2,
    rate_tol_pct: float = 0.05,
    rate_abs_tol: float = 0.0005,
    min_cells: int = 100,
    text_coverage_threshold: float = 0.3,
    cross_checks: list[Any] | None = None,
) -> VerifierResult:
    """Load golden reference, read submitted spreadsheet, and compare.

    Each task-specific verifier calls this with its own golden_path,
    concept_candidates, and tolerance settings.
    """
    # Load golden reference
    try:
        golden_sheet_names, golden_sheets_data = load_xlsx(golden_path)
    except FileNotFoundError as e:
        return VerifierResult(success=False, message=str(e))

    # ── Try direct download first (no tool server required) ──
    spreadsheet_id = extract_spreadsheet_id_from_messages(run_artifacts.messages)
    sheet_names: list[str] = []
    sheets_data: dict[str, list[list[Any]]] = {}

    if spreadsheet_id:
        logger.info("Extracted spreadsheet ID from messages: %s", spreadsheet_id)
        sheet_names, sheets_data, error = _download_xlsx_direct(spreadsheet_id)
        if error:
            logger.warning("Direct download failed (%s), falling back to tool server", error)

    # ── Fall back to tool server if direct download didn't work ──
    if not sheet_names:
        tool_server_url = run_artifacts.server_url("google-workspace-tool-server")
        if not tool_server_url:
            return VerifierResult(
                success=False,
                message="No spreadsheet ID in messages and no Google Workspace tool server URL",
            )

        client = ToolCallingClient(base_url=tool_server_url, request_timeout_s=30.0)
        try:
            spreadsheet_id, error = get_workspace_spreadsheet(client)
            if not spreadsheet_id:
                return VerifierResult(success=False, message=error)

            sheet_names, sheets_data, error = read_submitted_spreadsheet(client, spreadsheet_id)
            if error:
                return VerifierResult(success=False, message=error)
        finally:
            client.close()

    if not sheet_names:
        return VerifierResult(success=False, message="Spreadsheet has no sheets")

    try:
        # Run checks
        results = run_checks(
            sheet_names,
            sheets_data,
            golden_sheet_names,
            golden_sheets_data,
            concept_candidates,
            num_tol_pct=num_tol_pct,
            num_abs_tol=num_abs_tol,
            rate_tol_pct=rate_tol_pct,
            rate_abs_tol=rate_abs_tol,
            min_cells=min_cells,
            text_coverage_threshold=text_coverage_threshold,
            cross_checks=cross_checks,
        )

        total = len(results)
        passed = sum(1 for r in results if r["pass"])
        score = passed / total if total > 0 else 0.0

        return VerifierResult(
            success=score >= 0.8,
            message=f"{passed}/{total} checks passed ({score:.0%})",
            output=json.dumps(results, indent=2),
        )

    except Exception as e:
        logger.exception("Verifier error")
        return VerifierResult(success=False, message=f"Verifier error: {e}")
