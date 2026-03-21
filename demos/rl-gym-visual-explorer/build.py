#!/usr/bin/env python3
"""Build script for the Enterprise demo page.

Reads customer_service_package and hr_package data, transforms into the
Enterprise page's expected JSON format, and injects into index.html.
"""
import json, os, re, sys, glob

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(BASE, '..', '..')
CS_PKG = os.path.join(ROOT, 'customer_service_package')
HR_PKG = os.path.join(ROOT, 'hr_package')
FIN_PKG = os.path.join(ROOT, 'finance_package')
INDEX = os.path.join(BASE, 'index.html')

# ── Helpers ──

def read_json(path):
    with open(path, 'r') as f:
        return json.load(f)

def read_text(path):
    with open(path, 'r') as f:
        return f.read()

def messages_to_steps(messages):
    """Convert OpenAI-style messages array to Enterprise page steps format."""
    steps = []
    step_num = 0
    # Build a map of tool_call_id -> pending tool call info
    pending_tool_calls = {}

    for msg in messages:
        role = msg.get('role', '')
        content = msg.get('content', '')

        if role == 'assistant':
            # Handle text content (reasoning)
            if isinstance(content, str) and content and content.strip():
                step_num += 1
                steps.append({
                    's': step_num,
                    'type': 'reasoning',
                    'text': content.strip()
                })
            elif isinstance(content, list):
                # Content blocks format (Anthropic style)
                for block in content:
                    if isinstance(block, dict):
                        if block.get('type') == 'text' and block.get('text', '').strip():
                            step_num += 1
                            steps.append({
                                's': step_num,
                                'type': 'reasoning',
                                'text': block['text'].strip()
                            })
                        elif block.get('type') == 'tool_use':
                            step_num += 1
                            tc_id = block.get('id', '')
                            tool_name = block.get('name', 'unknown')
                            tool_input = block.get('input', {})
                            pending_tool_calls[tc_id] = step_num
                            steps.append({
                                's': step_num,
                                'type': 'tool_call',
                                'tool': tool_name,
                                'input': json.dumps(tool_input, ensure_ascii=False)[:1000],
                                'output': ''  # filled when tool result arrives
                            })

            # Handle tool_calls (OpenAI format)
            tool_calls = msg.get('tool_calls', [])
            for tc in tool_calls:
                step_num += 1
                fn = tc.get('function', {})
                tc_id = tc.get('id', tc.get('tool_call_id', ''))
                tool_name = fn.get('name', 'unknown')
                try:
                    tool_input = json.dumps(json.loads(fn.get('arguments', '{}')), ensure_ascii=False)[:2000]
                except:
                    tool_input = fn.get('arguments', '')[:1000]
                pending_tool_calls[tc_id] = step_num
                steps.append({
                    's': step_num,
                    'type': 'tool_call',
                    'tool': tool_name,
                    'input': tool_input,
                    'output': ''
                })

        elif role == 'tool':
            tc_id = msg.get('tool_call_id', '')
            output = ''
            if isinstance(content, str):
                output = content[:1500]
            elif isinstance(content, list):
                texts = [b.get('text', '') for b in content if isinstance(b, dict) and b.get('type') == 'text']
                output = '\n'.join(texts)[:1500]

            if tc_id in pending_tool_calls:
                sn = pending_tool_calls[tc_id]
                for step in steps:
                    if step['s'] == sn:
                        step['output'] = output
                        break

    return steps


def build_diffs(steps):
    """Generate diffs from tool call steps."""
    diffs = []
    for step in steps:
        if step['type'] != 'tool_call':
            continue
        tool = step.get('tool', '')
        label = tool.split('__')[-1] if '__' in tool else tool
        # Categorize
        if any(k in tool.lower() for k in ['email', 'send_email']):
            scope = 'email'
        elif any(k in tool.lower() for k in ['calendar', 'event', 'chronos']):
            scope = 'calendar'
        elif any(k in tool.lower() for k in ['chat', 'rocket', 'dm', 'send_dm']):
            scope = 'chat'
        elif any(k in tool.lower() for k in ['frappe', 'hris', 'hrms', 'helpdesk']):
            scope = 'hris'
        elif any(k in tool.lower() for k in ['sec-edgar', 'twelve-data', 'get_financials', 'get_filings', 'stock_price', 'price_history']):
            scope = 'data'
        elif any(k in tool.lower() for k in ['google-workspace', 'spreadsheet', 'sheet', 'drive', 'doc_content']):
            scope = 'workspace'
        else:
            scope = 'full'
        diffs.append({
            'step': step['s'],
            'scope': scope,
            'label': label.replace('_', ' '),
            'tool_name': tool,
            'summary': ''
        })
    return diffs


# ── Verifier check extraction ──

def extract_checks_from_py(py_path):
    """Extract individual check descriptions from a Python verifier file."""
    if not os.path.exists(py_path):
        return []
    code = read_text(py_path)
    checks = []
    # Match comments like: # 1a. ..., # A1. ..., # 1. ..., # A. ..., # B. ...
    for m in re.finditer(r'#\s+([A-Z]?\d*[a-z]?)\.\s+(.+)', code):
        check_id = m.group(1).strip()
        if not check_id:
            continue
        desc = m.group(2).strip().rstrip('—').strip()
        # Skip verdict/gate comments
        if any(kw in desc.lower() for kw in ['at least', 'hard gate', 'two conditions', 'final verdict']):
            continue
        checks.append(desc)
    return checks


def extract_failure_msgs(output_str, is_success=True):
    """Extract individual failure messages from a verifier output string."""
    failures = []

    # CS format: "X/Y checks passed ... Failures: msg1; msg2; ..."
    failures_match = re.search(r'Failures?:\s*(.+)$', output_str, re.DOTALL)
    if failures_match:
        for f in failures_match.group(1).strip().split(';'):
            f = f.strip()
            if f:
                failures.append(f)

    # NPC gate failure
    npc_match = re.search(r'NPC response gate failed.*?(\d+)/(\d+).*?responded.*?\(need (\d+)\+\)', output_str)
    if npc_match:
        failures.append(f'NPC response gate: only {npc_match.group(1)}/{npc_match.group(2)} internal NPCs responded (need {npc_match.group(3)}+)')

    # Check if X/Y shows partial pass (X < Y) — even if success=True
    count_match = re.search(r'(\d+)/(\d+)\s+checks?\s+passed', output_str)
    n_failed_from_count = 0
    if count_match:
        n_passed = int(count_match.group(1))
        n_total = int(count_match.group(2))
        n_failed_from_count = n_total - n_passed

    # HR format: if overall failed and no "Failures:" prefix, the entire output is failure(s)
    if not is_success and not failures:
        for f in output_str.split(';'):
            f = f.strip()
            if f and not re.match(r'^\d+/\d+\s+checks?\s+passed', f):
                failures.append(f)

    return failures, n_failed_from_count


STOP_WORDS = {'the', 'and', 'for', 'with', 'not', 'did', 'does', 'was', 'has', 'had',
               'are', 'were', 'been', 'being', 'have', 'from', 'that', 'this', 'but',
               'all', 'any', 'can', 'will', 'just', 'should', 'would', 'could', 'before',
               'after', 'into', 'about', 'than', 'then', 'each', 'only', 'check', 'found',
               'path', 'both', 'must', 'need'}

# Words that strongly indicate specific check types
ACTION_WORDS = {'event', 'calendar', 'email', 'chat', 'message', 'hris', 'update',
                'lookup', 'channel', 'sequencing', 'scope', 'discipline', 'probation',
                'meeting', 'schedule', 'escalation', 'escalated', 'ticket', 'triage',
                'responded', 'response', 'gate', 'npc', 'attendee', 'attendees'}

def match_failure_to_check(failure, checks):
    """Return the index of the check that best matches this failure, or -1."""
    fl = failure.lower()
    fail_words = set(re.findall(r'\w{3,}', fl)) - STOP_WORDS
    best_idx = -1
    best_score = 0
    for i, check in enumerate(checks):
        check_words = set(re.findall(r'\w{3,}', check.lower())) - STOP_WORDS
        overlap = check_words & fail_words
        if not overlap:
            continue
        # Score: count overlap, but weight action words 2x
        score = sum(2 if w in ACTION_WORDS else 1 for w in overlap)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= 1 else -1


def build_unified_verifier_data(all_checks, model_outputs, model_success=None):
    """Build pass/fail for each model using a unified check list.

    all_checks: list of check descriptions from .py file
    model_outputs: dict of {model_key: raw_verifier_output_string}
    model_success: dict of {model_key: bool} — whether verifier overall passed

    Returns: (unified_checks, {model_key: (passed_list, failed_list)})
    where both models have the same total = len(unified_checks)
    """
    if model_success is None:
        model_success = {}
    # Collect failures per model
    model_failures = {}
    model_fail_counts = {}
    for mk, output_str in model_outputs.items():
        is_success = model_success.get(mk, True)
        failures, n_failed_from_count = extract_failure_msgs(output_str, is_success=is_success)
        model_failures[mk] = failures
        model_fail_counts[mk] = n_failed_from_count

    # Build unified check list: start with .py checks, add unmatched failures
    unified = list(all_checks)
    for mk, failures in model_failures.items():
        for f in failures:
            idx = match_failure_to_check(f, unified)
            if idx == -1:
                unified.append(f)

    # For each model, determine pass/fail per unified check
    result = {}
    for mk, failures in model_failures.items():
        passed = []
        failed = []
        # Map each explicit failure to a unified check index
        failed_indices = set()
        for f in failures:
            idx = match_failure_to_check(f, unified)
            if idx >= 0:
                failed_indices.add(idx)

        # If we know the total fail count from X/Y but have fewer explicit failures,
        # mark additional checks as failed (from the end of the list)
        n_expected_fails = model_fail_counts[mk]
        if n_expected_fails > len(failed_indices):
            # Mark checks from the end that aren't already failed
            extra_needed = n_expected_fails - len(failed_indices)
            for i in reversed(range(len(unified))):
                if extra_needed <= 0:
                    break
                if i not in failed_indices:
                    failed_indices.add(i)
                    extra_needed -= 1

        seen_failures = set()
        for i, check in enumerate(unified):
            if i in failed_indices:
                matched_msg = check
                for f in failures:
                    if f not in seen_failures and match_failure_to_check(f, [check]) == 0:
                        matched_msg = f
                        seen_failures.add(f)
                        break
                failed.append(matched_msg)
            else:
                passed.append(check)
        result[mk] = (passed, failed)

    return unified, result


# ── Verifier file mapping ──

CS_VERIFIER_FILES = {
    206: 'customer_conversation_flow.py',
    207: 'escalation_flow.py',
    208: 'ticket_npc_interaction.py',
}

CS_BRIEFS = {
    206: 'The agent must de-escalate two frustrated customers simultaneously via RocketChat while following tier-appropriate SLA policies.',
    207: 'The agent must navigate a multi-level escalation chain to resolve an Enterprise customer\'s $25,000 credit demand with proper stakeholder coordination.',
    208: 'The agent must triage two simultaneous support issues by severity and customer tier, escalating a critical bug to engineering.',
}

HR_BRIEFS = {
    '145': 'The agent must handle escalated compliance training non-compliance across the operations team using HRIS, email, and calendar tools.',
    '161': 'The agent must coordinate cross-department workforce planning for an engineering team lead transition using HR systems and internal communications.',
    '79': 'The agent must conduct urgent wellbeing check-ins with the data science team following a teammate\'s sudden departure.',
}


# ── Customer Service ──

def build_cs_env():
    """Build the customer_service environment data."""
    profiles = read_json(os.path.join(CS_PKG, 'profiles.json'))
    tasks = {}

    for task_num in [206, 207, 208]:
        # Find task definition
        task_files = [f for f in os.listdir(CS_PKG) if f.startswith(str(task_num)) and f.endswith('.json')]
        if not task_files:
            continue
        task_def = read_json(os.path.join(CS_PKG, task_files[0]))

        # Find rubric markdown
        md_files = [f for f in os.listdir(CS_PKG) if f.startswith(str(task_num)) and f.endswith('.md')]
        rubric_md = read_text(os.path.join(CS_PKG, md_files[0])) if md_files else ''

        # Extract check names from verifier Python file
        verifier_py = CS_VERIFIER_FILES.get(task_num, '')
        all_checks = extract_checks_from_py(os.path.join(CS_PKG, verifier_py)) if verifier_py else []

        task_id = task_def.get('meta', {}).get('task_id', str(task_num))
        display_name = task_def.get('meta', {}).get('display_name', f'Task {task_num}')
        task_text = task_def.get('task', '')

        # First pass: collect raw data per model
        model_raw = {}
        for model_key, model_label in [('opus', 'Claude Opus 4.6'), ('grok', 'Grok')]:
            rollout_path = os.path.join(CS_PKG, f'rollout_{model_key}_{task_num}.json')
            if not os.path.exists(rollout_path):
                continue
            rollout = read_json(rollout_path)
            messages = rollout.get('live_output', {}).get('messages', [])
            result = rollout.get('result', {})
            verifiers = result.get('verifiers', [])

            raw_verifier_output = ''
            for v in verifiers:
                if 'universal_verifier' not in v.get('module', ''):
                    raw_verifier_output = v.get('output', '')
                    break

            rubric_output = ''
            for v in verifiers:
                if 'universal_verifier' in v.get('module', ''):
                    rubric_output = v.get('output', '')
                    break

            model_raw[model_key] = {
                'label': model_label, 'rollout': rollout, 'messages': messages,
                'result': result, 'raw_verifier': raw_verifier_output,
                'rubric_output': rubric_output
            }

        if not model_raw:
            continue

        # Build unified verifier checks across all models
        raw_outputs = {mk: d['raw_verifier'] for mk, d in model_raw.items()}
        success_map = {}
        for mk, d in model_raw.items():
            verifiers = d['result'].get('verifiers', [])
            success_map[mk] = all(v.get('success', True) for v in verifiers
                                  if 'universal_verifier' not in v.get('module', ''))
        _, verifier_results = build_unified_verifier_data(all_checks, raw_outputs, success_map)

        # Second pass: build model objects
        models = {}
        for model_key, raw in model_raw.items():
            rollout = raw['rollout']
            result = raw['result']
            steps = messages_to_steps(raw['messages'])
            diffs = build_diffs(steps)
            metrics = result.get('metrics', {})

            passed, failed = verifier_results[model_key]
            verifier_output = f"passed={passed}\nmissing={failed}"

            models[model_key] = {
                'meta': {
                    'rollout_id': rollout.get('rollout_id', ''),
                    'task_id': rollout.get('task_id', task_id),
                    'model': raw['label'],
                    'status': rollout.get('status', 'completed'),
                    'steps_taken': result.get('steps_taken', rollout.get('current_step', len(steps))),
                    'max_steps': rollout.get('max_steps', 100),
                    'duration_seconds': result.get('duration_seconds', 0),
                    'cost_usd': round(result.get('estimated_cost_usd', 0), 4),
                    'prompt_tokens': metrics.get('prompt_tokens_total', 0),
                    'completion_tokens': metrics.get('completion_tokens_total', 0),
                    'composite_score': result.get('composite_score', result.get('reward', 0)),
                    'success': result.get('success', rollout.get('status') == 'completed')
                },
                'steps': steps,
                'diffs': diffs,
                'final_observation': '',
                'verifier_output': verifier_output,
                'rubric_output': raw['rubric_output']
            }

        if models:
            tasks[task_num] = {
                'task_id': task_id,
                'display_name': display_name,
                'brief': CS_BRIEFS.get(task_num, ''),
                'task': task_text,
                'rubric_definition': rubric_md,
                'seed_data': {},
                'npc_profiles': profiles,
                'models': models
            }

    return tasks


# ── HR ──

def build_hr_env():
    """Build the HR environment data from hr_package."""
    tasks_dir = os.path.join(HR_PKG, 'eval_tasks_package')
    traj_dir = os.path.join(HR_PKG, 'eval-trajectories')

    # Load NPC profiles
    profiles_path = os.path.join(tasks_dir, 'Seeddata', 'npcs', 'profiles.json')
    profiles = read_json(profiles_path) if os.path.exists(profiles_path) else []

    # Load seed data (policies)
    seed_data = {}
    seed_dir = os.path.join(tasks_dir, 'Seeddata', 'seed_data')
    if os.path.exists(seed_dir):
        for fname in sorted(os.listdir(seed_dir)):
            fpath = os.path.join(seed_dir, fname)
            if os.path.isfile(fpath) and not fname.startswith('.'):
                seed_data[fname] = read_text(fpath)

    # Load org chart from seed data
    org_chart_path = os.path.join(seed_dir, 'org_chart.md')
    if os.path.exists(org_chart_path):
        seed_data['org_chart.md'] = read_text(org_chart_path)

    # Find task directories — pick 3 tasks with strongest opus vs grok contrast
    HR_SELECTED = {'161', '79', '145'}  # workforce planning (+0.47), wellbeing check-ins (+0.39), compliance training (+0.38)
    task_dirs = []
    for d in sorted(os.listdir(tasks_dir)):
        dp = os.path.join(tasks_dir, d)
        if os.path.isdir(dp) and d not in ('Seeddata', '__pycache__') and not d.startswith('.'):
            num_match = re.match(r'^(\d+)', d)
            if num_match and num_match.group(1) in HR_SELECTED:
                task_dirs.append(d)

    # Pre-extract check definitions from all verifier .py files
    hr_check_defs = {}
    for td in task_dirs:
        td_path = os.path.join(tasks_dir, td)
        py_files = [f for f in os.listdir(td_path) if f.startswith('v') and f.endswith('.py')]
        if py_files:
            num_match = re.match(r'^(\d+)', td)
            if num_match:
                checks = extract_checks_from_py(os.path.join(td_path, py_files[0]))
                hr_check_defs[num_match.group(1)] = checks

    tasks = {}
    for td in task_dirs:
        td_path = os.path.join(tasks_dir, td)
        # Find JSON task definition
        json_files = [f for f in os.listdir(td_path) if f.endswith('.json') and not f.startswith('.')]
        if not json_files:
            continue
        task_def = read_json(os.path.join(td_path, json_files[0]))
        meta = task_def.get('meta', {})
        task_id = meta.get('task_id', td)
        display_name = meta.get('display_name', td)
        task_text = task_def.get('task', '')

        # Find rubric markdown
        md_files = [f for f in os.listdir(td_path) if f.endswith('.md')]
        rubric_md = ''
        if md_files:
            rubric_md = read_text(os.path.join(td_path, md_files[0]))

        # Extract numeric task id for matching trajectories
        num_match = re.match(r'^(\d+)', td)
        task_num = num_match.group(1) if num_match else td

        # First pass: collect raw data per model
        model_raw = {}
        for model_dir, model_key in [('claude-opus-4.6-eval', 'opus'), ('grok-4.20-beta-eval', 'grok')]:
            model_path = os.path.join(traj_dir, model_dir)
            if not os.path.exists(model_path):
                continue
            pattern = f'trajectory-{task_num}_*.json'
            matches = glob.glob(os.path.join(model_path, pattern))
            main_runs = [m for m in matches if not re.search(r'-r\d+\.json$', m)]
            if not main_runs:
                main_runs = matches[:1]
            if not main_runs:
                continue

            traj = read_json(main_runs[0])
            result = traj.get('result', {})
            artifacts = traj.get('artifacts', {})
            messages = artifacts.get('messages', [])
            verifiers = result.get('verifiers', [])
            raw_output = '; '.join(v.get('output', '') for v in verifiers if v.get('output'))

            judge = result.get('judge', {})
            rubric_output = json.dumps(judge, ensure_ascii=False) if judge else ''

            model_raw[model_key] = {
                'traj': traj, 'result': result, 'messages': messages,
                'raw_verifier': raw_output, 'rubric_output': rubric_output
            }

        if not model_raw:
            continue

        # Build unified verifier checks across all models
        all_checks = hr_check_defs.get(task_num, [])
        raw_outputs = {mk: d['raw_verifier'] for mk, d in model_raw.items()}
        success_map = {mk: all(v.get('success', True) for v in d['result'].get('verifiers', []))
                       for mk, d in model_raw.items()}
        _, verifier_results = build_unified_verifier_data(all_checks, raw_outputs, success_map)

        # Second pass: build model objects
        models = {}
        for model_key, raw in model_raw.items():
            traj = raw['traj']
            result = raw['result']
            steps = messages_to_steps(raw['messages'])
            diffs = build_diffs(steps)
            metrics = result.get('metrics', {})

            passed, failed = verifier_results[model_key]
            verifier_output = f"passed={passed}\nmissing={failed}"

            models[model_key] = {
                'meta': {
                    'rollout_id': traj.get('rollout_id', ''),
                    'task_id': traj.get('task_id', task_id),
                    'model': 'Claude Opus 4.6' if model_key == 'opus' else 'Grok 4.20',
                    'status': traj.get('status', 'completed'),
                    'steps_taken': result.get('steps_taken', len(steps)),
                    'max_steps': traj.get('max_steps', 100),
                    'duration_seconds': result.get('duration_seconds', 0),
                    'cost_usd': round(result.get('estimated_cost_usd', 0), 4),
                    'prompt_tokens': metrics.get('prompt_tokens_total', 0),
                    'completion_tokens': metrics.get('completion_tokens_total', 0),
                    'composite_score': result.get('reward', 0),
                    'success': result.get('success', True)
                },
                'steps': steps,
                'diffs': diffs,
                'final_observation': '',
                'verifier_output': verifier_output,
                'rubric_output': raw['rubric_output']
            }

        if models:
            tasks[task_num] = {
                'task_id': task_id,
                'display_name': display_name,
                'brief': HR_BRIEFS.get(task_num, ''),
                'task': task_text,
                'rubric_definition': rubric_md,
                'models': models
            }

    return tasks, seed_data, profiles


# ── Finance ──

# Split points for instruction.md: line number where "expected output" section starts
_FIN_INSTRUCTION_SPLITS = {
    'finance_task_1': '## Documentation Standards',
    'finance_task_2': '## Critical Instructions',
}

FINANCE_TASKS = {
    'finance_task_1': {
        'display_name': 'Coca-Cola Investment Analysis',
        'task_id': 'coca-cola-investment-analysis',
        'brief': 'The agent must build an Investment Manager–ready spreadsheet analyzing Coca-Cola (KO) using FY2022–2024 SEC filings and market data.',
    },
    'finance_task_2': {
        'display_name': 'Trade Reconciliation',
        'task_id': 'trade-reconciliation',
        'brief': 'The agent must reconcile an internal trade blotter against broker confirmations, classify all discrepancies, and quantify dollar impact.',
    },
}


def _split_instruction(text, split_header):
    """Split instruction.md into task description and expected output formatting."""
    idx = text.find(split_header)
    if idx < 0:
        return text, ''
    return text[:idx].rstrip(), text[idx:].strip()


def _build_unified_finance_checks(model_raw):
    """Build a unified check list across all models for finance verifiers.

    Finance verifier output is JSON: [{criteria: str, pass: bool}, ...]
    Returns: (unified_checks, {model_key: (passed_list, failed_list)})
    where both models show the same total count.
    """
    # Parse checks per model
    model_checks = {}
    for mk, raw in model_raw.items():
        try:
            checks = json.loads(raw['raw_verifier'])
            model_checks[mk] = {c['criteria']: c['pass'] for c in checks}
        except Exception:
            model_checks[mk] = {}

    # Build unified check list (union of all criteria, preserving order from first model)
    unified = []
    seen = set()
    for mk in model_checks:
        for criteria in model_checks[mk]:
            if criteria not in seen:
                unified.append(criteria)
                seen.add(criteria)

    # For each model, determine pass/fail per unified check
    result = {}
    for mk in model_checks:
        passed = []
        failed = []
        for criteria in unified:
            if criteria in model_checks[mk]:
                if model_checks[mk][criteria]:
                    passed.append(criteria)
                else:
                    failed.append(criteria)
            else:
                # Check not in this model's output — mark as failed
                failed.append(criteria)
        result[mk] = (passed, failed)

    return unified, result


def _read_xlsx_data(path, max_rows=100):
    """Read an xlsx file into a JSON-friendly structure."""
    try:
        import openpyxl
    except ImportError:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for r in range(1, min(max_rows + 1, ws.max_row + 1)):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    row.append('')
                elif isinstance(v, float):
                    row.append(round(v, 6))
                else:
                    row.append(str(v))
            rows.append(row)
        result[sn] = rows
    return result


def _read_input_files(task_path):
    """Read all xlsx input files for a finance task as table data."""
    input_dir = os.path.join(task_path, 'input_file')
    files = {}
    if os.path.exists(input_dir):
        for fname in sorted(os.listdir(input_dir)):
            fpath = os.path.join(input_dir, fname)
            if os.path.isfile(fpath) and fname.endswith('.xlsx'):
                files[fname] = _read_xlsx_data(fpath)
    return files


def _read_output_xlsx(task_path, model_dir):
    """Read output xlsx for a model run."""
    model_path = os.path.join(task_path, model_dir)
    for fname in os.listdir(model_path):
        if fname.endswith('.xlsx'):
            return _read_xlsx_data(os.path.join(model_path, fname))
    return {}


def build_finance_env():
    """Build the finance environment data from finance_package."""
    tasks = {}
    seed_data = {}

    for task_dir_name, task_meta in FINANCE_TASKS.items():
        task_path = os.path.join(FIN_PKG, task_dir_name)
        if not os.path.exists(task_path):
            continue

        # Read instruction and rubric
        full_instruction = read_text(os.path.join(task_path, 'instruction.md'))
        rubric_md = read_text(os.path.join(task_path, 'rubric.md'))

        # Split instruction into task description and expected output formatting
        split_header = _FIN_INSTRUCTION_SPLITS.get(task_dir_name, '')
        task_desc, output_format = _split_instruction(full_instruction, split_header)

        # Read xlsx input files as table data
        input_xlsx = _read_input_files(task_path)

        # Read xlsx output files per model
        model_xlsx = {}
        for model_dir, model_key in [('opus-4.6-output', 'opus'), ('grok-4.20-output', 'grok')]:
            model_xlsx[model_key] = _read_output_xlsx(task_path, model_dir)

        # Build models
        model_raw = {}
        for model_dir, model_key, model_label in [
            ('opus-4.6-output', 'opus', 'Claude Opus 4.6'),
            ('grok-4.20-output', 'grok', 'Grok 4.20'),
        ]:
            model_path = os.path.join(task_path, model_dir)
            if not os.path.exists(model_path):
                continue

            # Find trajectory JSON (exclude rubric_score.json)
            traj_files = [f for f in os.listdir(model_path)
                          if f.endswith('.json') and f != 'rubric_score.json']
            if not traj_files:
                continue

            traj = read_json(os.path.join(model_path, traj_files[0]))
            result = traj.get('result', {})
            messages = traj.get('live_output', {}).get('messages', [])

            # Verifier output — JSON array of {criteria, pass}
            verifiers = result.get('verifiers', [])
            raw_verifier = ''
            for v in verifiers:
                if 'universal_verifier' not in v.get('module', ''):
                    raw_verifier = v.get('output', '')
                    break

            # Rubric output from rubric_score.json
            rubric_score_path = os.path.join(model_path, 'rubric_score.json')
            rubric_output = read_text(rubric_score_path) if os.path.exists(rubric_score_path) else ''

            model_raw[model_key] = {
                'label': model_label,
                'traj': traj,
                'result': result,
                'messages': messages,
                'raw_verifier': raw_verifier,
                'rubric_output': rubric_output,
            }

        if not model_raw:
            continue

        # Build unified verifier checks across all models
        _, verifier_results = _build_unified_finance_checks(model_raw)

        # Build model objects
        models = {}
        for model_key, raw in model_raw.items():
            traj = raw['traj']
            result = raw['result']
            steps = messages_to_steps(raw['messages'])
            diffs = build_diffs(steps)
            metrics = result.get('metrics', {})

            passed, failed = verifier_results.get(model_key, ([], []))
            verifier_output = f"passed={passed}\nmissing={failed}"

            models[model_key] = {
                'meta': {
                    'rollout_id': traj.get('rollout_id', ''),
                    'task_id': traj.get('task_id', task_meta['task_id']),
                    'model': raw['label'],
                    'status': traj.get('status', 'completed'),
                    'steps_taken': result.get('steps_taken', len(steps)),
                    'max_steps': traj.get('max_steps', 100),
                    'duration_seconds': result.get('duration_seconds', 0),
                    'cost_usd': round(result.get('estimated_cost_usd', 0), 4),
                    'prompt_tokens': metrics.get('prompt_tokens_total', 0),
                    'completion_tokens': metrics.get('completion_tokens_total', 0),
                    'composite_score': result.get('composite_score', result.get('reward', 0)),
                    'success': result.get('success', True),
                },
                'steps': steps,
                'diffs': diffs,
                'final_observation': '',
                'verifier_output': verifier_output,
                'rubric_output': raw['rubric_output'],
                'output_xlsx': model_xlsx.get(model_key, {}),
            }

        if models:
            tasks[task_dir_name] = {
                'task_id': task_meta['task_id'],
                'display_name': task_meta['display_name'],
                'brief': task_meta.get('brief', ''),
                'task': task_desc,
                'output_format': output_format,
                'rubric_definition': rubric_md,
                'input_xlsx': input_xlsx,
                'models': models,
            }

    return tasks, seed_data


def build_cs_policies():
    """Generate fictional customer service policies and documentation."""
    return {
        "escalation_policy.md": """# Escalation Policy — Weaver Enterprises Inc.

## Purpose
This policy defines the escalation framework for customer support issues that exceed frontline resolution authority or require cross-functional coordination.

## Escalation Levels

### Level 1 — Frontline Agent
- Handle standard inquiries, billing questions, and known-issue troubleshooting
- Authority: issue refunds up to $500, apply account credits up to $200
- Escalate if: customer requests manager, issue involves >$500, or outage affecting enterprise tier

### Level 2 — Support Manager (jennifer.lee)
- Review escalated cases, approve credits up to $5,000
- Authority: waive fees, extend trial periods, expedite engineering tickets
- Escalate if: customer threatens contract termination, credit request >$5,000, or legal/compliance risk

### Level 3 — Billing Specialist (diana.walsh)
- Consult on financial impact, account history, revenue at risk
- Authority: assess policy-appropriate credit amounts, review account spend history
- Provide recommendations to VP for large credits

### Level 4 — VP of Customer Success (robert.hayes)
- Final decision authority on credits >$5,000
- Authority: approve service credits up to $50,000, negotiate contract amendments
- Must receive synthesized briefing from both Manager and Billing before decision

## Escalation Order
Agents MUST follow the chain: **Agent → Manager → Billing → VP**. Skipping levels is prohibited except in cases involving immediate safety or legal risk.

## Documentation Requirements
All escalations must be documented in a Helpdesk ticket with:
- Customer name, tier, and contact history
- Issue summary and business impact
- Each escalation step and stakeholder response
- Final resolution and follow-up actions
""",

        "customer_tier_policy.md": """# Customer Tier Policy

## Tier Definitions

### Enterprise Tier
- Annual contract value: >$50,000
- Dedicated account manager
- Priority SLA: 1-hour initial response, 4-hour resolution target
- Direct escalation path to engineering
- Quarterly business reviews

### SMB (Small & Medium Business) Tier
- Annual contract value: $5,000–$50,000
- Shared account management
- Standard SLA: 4-hour initial response, 24-hour resolution target
- Standard support queue

### Individual Tier
- Monthly subscription or pay-as-you-go
- Self-service support with community forums
- Standard SLA: 24-hour initial response, 72-hour resolution target

## Tier-Specific Communication Guidelines

**Enterprise customers** require:
- Urgent, professional tone acknowledging business impact
- Proactive status updates every 2 hours during incidents
- Named point of contact throughout resolution

**SMB customers** require:
- Empathetic, patient communication
- De-escalation techniques for frustrated customers
- Knowledge base article references where applicable

**Individual customers** require:
- Friendly, clear instructions
- Self-service resource links
- Upgrade path suggestions when appropriate
""",

        "sla_policy.md": """# Service Level Agreement (SLA) Policy

## Response Time Targets

| Priority | Enterprise | SMB | Individual |
|----------|-----------|-----|------------|
| Critical (P1) | 15 min | 1 hour | 4 hours |
| High (P2) | 1 hour | 4 hours | 24 hours |
| Medium (P3) | 4 hours | 24 hours | 48 hours |
| Low (P4) | 24 hours | 48 hours | 72 hours |

## Priority Classification

- **Critical (P1)**: Service fully unavailable, data loss risk, security breach
- **High (P2)**: Major feature broken, significant performance degradation, blocking operations
- **Medium (P3)**: Feature partially impaired, workaround available
- **Low (P4)**: Cosmetic issues, feature requests, general inquiries

## SLA Breach Procedures
1. Automatic notification to Support Manager at 75% of SLA window
2. Escalation to VP of Customer Success at SLA breach
3. Post-incident review required for all Enterprise P1/P2 breaches
4. Service credits automatically applied per contract terms

## Service Credit Schedule
- P1 SLA breach: 5% monthly credit per incident
- P2 SLA breach: 2% monthly credit per incident
- Cumulative cap: 30% of monthly contract value
""",

        "knowledge_base_guide.md": """# Knowledge Base Usage Guide

## Overview
The Weaver Enterprises knowledge base contains troubleshooting articles, FAQs, and product documentation that agents should reference during customer interactions.

## Search Best Practices
1. Search by error code or symptom keywords before escalating
2. Reference article titles in customer communications
3. Link relevant articles in Helpdesk ticket comments

## Key Article Categories

### Billing & Account
- Refund Policy and Procedures
- Billing FAQ and Common Issues
- Account Upgrade/Downgrade Process

### Technical Troubleshooting
- Performance Troubleshooting Guide
- API Error Code Reference
- Bulk Import Best Practices
- Dashboard Loading Optimization

### Platform & Features
- Getting Started Guide
- Integration Setup Documentation
- Data Export and Backup Procedures

## Article Update Process
- Support agents may flag outdated articles via #kb-updates channel
- Product team reviews and updates articles quarterly
- All customer-facing articles require review before publication
""",

        "org_chart.md": """# Weaver Enterprises Inc. — Support Organization

## Leadership
- **Robert Hayes** — VP of Customer Success
  - **Jennifer Lee** — Support Manager
    - **Sarah Johnson** — Senior Support Agent
    - **Mike Williams** — Support Agent (Billing Specialist)
    - **Rachel Green** — Support Agent (Technical Specialist)
    - **Tom Nguyen** — Support Agent (Junior)

## Account Management
- **Amanda Reeves** — Senior Account Manager (Enterprise Accounts)
- **Carlos Mendez** — Account Manager (SMB Accounts)

## Billing
- **James Foster** — Billing Manager
  - **Diana Walsh** — Billing Specialist

## Engineering
- **Sandra Kim** — Director of Engineering
  - **Marcus Chen** — Engineering Lead
    - **Priya Sharma** — Senior Backend Engineer (API Infrastructure)
    - **Kevin Zhang** — DevOps Engineer

## Quality & Product
- **Lisa Tanaka** — QA Lead
- **Ryan Brooks** — Product Manager
- **Emma Davis** — UX Designer

## Tooling
- **RocketChat** — Internal messaging and customer DMs
- **Frappe Helpdesk** — Ticket management and knowledge base
""",

        "communication_guidelines.md": """# Customer Communication Guidelines

## De-escalation Framework
When handling frustrated or angry customers:

1. **Acknowledge** — Validate their frustration: "I understand this is frustrating"
2. **Apologize** — Take ownership: "I'm sorry for the inconvenience"
3. **Act** — State what you will do: "Let me look into this right away"
4. **Follow up** — Set expectations: "I'll update you within the hour"

## Tone by Situation

### Frustrated Customer (repeat issues)
- Lead with empathy, not defensiveness
- Reference their history to show you've reviewed the account
- Avoid phrases: "As I mentioned", "Per our policy", "Unfortunately"
- Use phrases: "I completely understand", "Let me personally make sure", "You deserve better"

### Enterprise Customer (business impact)
- Treat every issue as urgent
- Quantify the impact acknowledgment: "I understand this is blocking your team"
- Provide concrete timelines, not vague assurances
- Loop in account manager for visibility

### Technical Issues
- Gather specific details: error messages, timestamps, reproduction steps
- Avoid jargon with non-technical customers
- Provide clear next steps and expected resolution timeline

## Concurrent Customer Handling
- Never sacrifice quality for speed
- Interleave responses across active conversations
- Prioritize by tier and severity
- Document all conversations in the ticketing system
""",
    }


def main():
    print("Building Enterprise demo page...")

    # Read existing page
    html = read_text(INDEX)

    # Extract existing finance data
    edata_match = re.search(r'<script id="edata" type="application/json">([\s\S]*?)</script>', html)
    if not edata_match:
        print("ERROR: Could not find edata in index.html")
        sys.exit(1)

    # Build new environments
    cs_tasks = build_cs_env()
    hr_tasks, hr_seed_data, hr_profiles = build_hr_env()
    fin_tasks, fin_seed_data = build_finance_env()

    print(f"  Customer Service: {len(cs_tasks)} tasks")
    for num, t in cs_tasks.items():
        print(f"    Task {num}: {t['display_name']} ({len(t['models'])} models)")
    print(f"  HR: {len(hr_tasks)} tasks")
    for num, t in hr_tasks.items():
        print(f"    Task {num}: {t['display_name'][:60]}... ({len(t['models'])} models)")
    print(f"  Finance: {len(fin_tasks)} tasks")
    for num, t in fin_tasks.items():
        print(f"    {num}: {t['display_name']} ({len(t['models'])} models)")

    # For environments with multiple tasks, we need a different data structure.
    # The page expects: DATA[env] = {task, models, ...}
    # We'll use: DATA[env] = {tasks: {id: {task, models, ...}}, current_task: first_id, ...}
    # And for backward compat, also set top-level task/models from current_task.

    # Build the new data object
    new_data = {}

    # Customer Service: use multi-task format
    cs_seed_data = build_cs_policies()
    if cs_tasks:
        first_key = list(cs_tasks.keys())[0]
        first_task = cs_tasks[first_key]
        new_data['customer_service'] = {
            'task': first_task['task'],
            'rubric_definition': first_task['rubric_definition'],
            'seed_data': cs_seed_data,
            'npc_profiles': first_task.get('npc_profiles', []),
            'models': first_task['models'],
            'tasks': {str(k): v for k, v in cs_tasks.items()},
            'current_task': str(first_key)
        }

    # HR: use multi-task format
    if hr_tasks:
        first_key = list(hr_tasks.keys())[0]
        first_task = hr_tasks[first_key]
        new_data['hr'] = {
            'task': first_task['task'],
            'rubric_definition': first_task['rubric_definition'],
            'seed_data': hr_seed_data,
            'npc_profiles': hr_profiles,
            'models': first_task['models'],
            'tasks': {str(k): v for k, v in hr_tasks.items()},
            'current_task': str(first_key)
        }

    # Finance: use multi-task format from finance_package
    if fin_tasks:
        first_key = list(fin_tasks.keys())[0]
        first_task = fin_tasks[first_key]
        new_data['finance'] = {
            'task': first_task['task'],
            'rubric_definition': first_task['rubric_definition'],
            'seed_data': fin_seed_data,
            'npc_profiles': [],
            'models': first_task['models'],
            'tasks': {str(k): v for k, v in fin_tasks.items()},
            'current_task': str(first_key)
        }

    # Serialize
    data_json = json.dumps(new_data, ensure_ascii=False, separators=(',', ':'))

    # Replace edata in HTML
    new_html = html[:edata_match.start(1)] + data_json + html[edata_match.end(1):]

    # Update environment pills: Finance first (active), then HR, then CS — as vertical list
    old_pills = '''<button class="env-pill active" onclick="switchEnv('hr')">HR</button>
    <button class="env-pill" onclick="switchEnv('finance')">Finance</button>'''
    new_pills = '''<button class="env-pill active" onclick="switchEnv('finance')">Finance</button>
    <button class="env-pill" onclick="switchEnv('hr')">HR</button>
    <button class="env-pill" onclick="switchEnv('customer_service')">Customer Service</button>'''
    new_html = new_html.replace(old_pills, new_pills)

    # Update model name mappings
    old_mn = "const MN={gpt_5_2:'GPT 5.2',gemini_3_pro:'Gemini 3 Pro'};"
    new_mn = "const MN={gpt_5_2:'GPT 5.2',gemini_3_pro:'Gemini 3 Pro',opus:'Claude Opus 4.6',grok:'Grok 4.20'};"
    new_html = new_html.replace(old_mn, new_mn)

    # Update default env and model
    old_init = "let DATA=null, currentEnv='hr', currentModel='gpt_5_2', currentView='rollouts';"
    new_init = "let DATA=null, currentEnv='finance', currentModel='opus', currentView='rollouts';"
    new_html = new_html.replace(old_init, new_init)

    # Update switchEnv to handle env pill matching (customer_service vs display text)
    old_switch = "document.querySelectorAll('.env-pill').forEach(p=>p.classList.toggle('active',p.textContent.trim().toLowerCase()===env));"
    new_switch = "document.querySelectorAll('.env-pill').forEach(p=>p.classList.toggle('active',p.dataset.env===env));"
    new_html = new_html.replace(old_switch, new_switch)

    # Add data-env attributes to pills
    new_html = new_html.replace(
        '''<button class="env-pill active" onclick="switchEnv('finance')">Finance</button>''',
        '''<button class="env-pill active" data-env="finance" onclick="switchEnv('finance')">Finance</button>'''
    )
    new_html = new_html.replace(
        '''<button class="env-pill" onclick="switchEnv('hr')">HR</button>''',
        '''<button class="env-pill" data-env="hr" onclick="switchEnv('hr')">HR</button>'''
    )
    new_html = new_html.replace(
        '''<button class="env-pill" onclick="switchEnv('customer_service')">Customer Service</button>''',
        '''<button class="env-pill" data-env="customer_service" onclick="switchEnv('customer_service')">Customer Service</button>'''
    )

    # Add task selection UI and logic
    # We need to add a task selector that shows when an env has multiple tasks.
    # Insert task pills section after the env-pills div
    task_pills_html = '''  <div class="sidebar-divider" id="task-divider" style="display:none"></div>
  <div class="sidebar-section" id="task-section-label" style="display:none">Task</div>
  <div class="task-pills" id="task-pills" style="display:none;padding:4px 12px;max-height:300px;overflow-y:auto"></div>'''

    new_html = new_html.replace(
        '  <div class="sidebar-divider"></div>\n  <nav class="sidebar-nav">',
        task_pills_html + '\n  <div class="sidebar-divider"></div>\n  <nav class="sidebar-nav">'
    )

    # Add env-pill vertical list + task pill CSS + seed-table CSS
    extra_css = """
.env-pills { flex-direction:column !important; gap:6px !important; padding:8px 16px !important; }
.env-pill { display:block !important; width:100% !important; text-align:left !important; padding:8px 12px !important; border-radius:8px !important; font-size:12.5px !important; font-weight:500 !important; border:1.5px solid var(--g200) !important; background:var(--g50) !important; color:var(--g700) !important; cursor:pointer; transition:all .15s; }
.env-pill:hover { border-color:var(--orange) !important; background:var(--orange-light) !important; color:var(--orange-dark) !important; }
.env-pill.active { background:var(--orange) !important; border-color:var(--orange) !important; color:white !important; font-weight:600 !important; }
html.dark .env-pill { border-color:#444 !important; background:#1e1e1e !important; color:#bbb !important; }
html.dark .env-pill:hover { border-color:var(--orange) !important; background:#2a2018 !important; }
html.dark .env-pill.active { background:var(--orange) !important; border-color:var(--orange) !important; color:white !important; }
.task-pills { display:flex; flex-direction:column; gap:6px; padding:8px 16px !important; }
.task-pill { padding:9px 12px; border-radius:8px; font-size:12px; font-weight:500; cursor:pointer; border:1.5px solid var(--g200); background:var(--g50); color:var(--g700); text-align:left; transition:all .15s; line-height:1.4; }
.task-pill:hover { border-color:var(--orange); background:var(--orange-light); color:var(--orange-dark); }
.task-pill.active { background:var(--orange); border-color:var(--orange); color:white; font-weight:600; }
html.dark .task-pill { border-color:#444; background:#1e1e1e; color:#bbb; }
html.dark .task-pill:hover { border-color:var(--orange); background:#2a2018; color:var(--orange); }
html.dark .task-pill.active { background:var(--orange); border-color:var(--orange); color:white; }
.seed-table { border-collapse:collapse; width:100%; font-size:12px; }
.seed-table th { background:var(--g100); position:sticky; top:0; z-index:1; font-weight:600; font-size:11px; color:var(--g500); padding:6px 10px; text-align:left; border:1px solid var(--g200); }
.seed-table td { padding:4px 10px; border:1px solid var(--g200); max-width:300px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.seed-table tr:hover td { background:var(--orange-light); }
.seed-table .num { text-align:right; font-family:monospace; }
html.dark .seed-table th { background:#2a2a2a; color:#aaa; border-color:#444; }
html.dark .seed-table td { border-color:#444; }
html.dark .seed-table tr:hover td { background:#2a2018; }
"""
    new_html = new_html.replace(
        '</style>',
        extra_css + '</style>',
        1  # first occurrence only
    )

    # Add task switching JS - insert before the init() function definition
    task_js = r"""
function renderTaskPills(){
  var env=DATA[currentEnv];
  var pills=document.getElementById('task-pills');
  var divider=document.getElementById('task-divider');
  var label=document.getElementById('task-section-label');
  if(!env.tasks||Object.keys(env.tasks).length<=1){
    pills.style.display='none';divider.style.display='none';label.style.display='none';return;
  }
  pills.style.display='flex';divider.style.display='';label.style.display='';
  var ct=env.current_task||Object.keys(env.tasks)[0];
  pills.innerHTML=Object.entries(env.tasks).map(function(e){
    var k=e[0],t=e[1];
    return '<button class="task-pill'+(k===ct?' active':'')+'" data-task="'+k+'" onclick="switchTask(\''+k+'\')" title="'+t.display_name.replace(/"/g,'&quot;')+'">'+t.display_name+'</button>';
  }).join('');
}
function switchTask(taskKey){
  var env=DATA[currentEnv];
  if(!env.tasks||!env.tasks[taskKey])return;
  env.current_task=taskKey;
  var t=env.tasks[taskKey];
  env.task=t.task;env.rubric_definition=t.rubric_definition;
  env.models=t.models;env.npc_profiles=t.npc_profiles||env.npc_profiles;
  currentModel=Object.keys(t.models)[0];
  renderTaskPills();renderModelTabs();
  if(currentView==='rollouts')renderRollouts();else renderEnvironment();
}
"""
    new_html = new_html.replace(
        'function switchEnv(env){',
        task_js + 'function switchEnv(env){'
    )

    # Update switchEnv to also render task pills and select first task
    old_switch_fn = """function switchEnv(env){
  currentEnv=env;
  document.querySelectorAll('.env-pill').forEach(p=>p.classList.toggle('active',p.dataset.env===env));
  currentModel=Object.keys(DATA[env].models)[0];
  renderModelTabs();
  if(currentView==='rollouts')renderRollouts();else renderEnvironment();
}"""
    new_switch_fn = """function switchEnv(env){
  currentEnv=env;
  document.querySelectorAll('.env-pill').forEach(p=>p.classList.toggle('active',p.dataset.env===env));
  var e=DATA[env];
  if(e.tasks&&Object.keys(e.tasks).length>0){
    var ct=e.current_task||Object.keys(e.tasks)[0];
    e.current_task=ct;var t=e.tasks[ct];
    e.task=t.task;e.rubric_definition=t.rubric_definition;
    e.models=t.models;e.npc_profiles=t.npc_profiles||e.npc_profiles;
  }
  currentModel=Object.keys(e.models)[0];
  renderTaskPills();renderModelTabs();
  if(currentView==='rollouts')renderRollouts();else renderEnvironment();
}"""
    new_html = new_html.replace(old_switch_fn, new_switch_fn)

    # Update init to call renderTaskPills and switchTask for initial load
    old_init_call = "DATA=JSON.parse(document.getElementById('edata').textContent);init();"
    new_init_call = """DATA=JSON.parse(document.getElementById('edata').textContent);
// Initialize first task for multi-task envs
Object.keys(DATA).forEach(function(ek){
  var e=DATA[ek];
  if(e.tasks&&Object.keys(e.tasks).length>0){
    var ct=e.current_task||Object.keys(e.tasks)[0];
    e.current_task=ct;var t=e.tasks[ct];
    e.task=t.task;e.rubric_definition=t.rubric_definition;
    e.models=t.models;if(t.npc_profiles)e.npc_profiles=t.npc_profiles;
  }
});
init();renderTaskPills();"""
    new_html = new_html.replace(old_init_call, new_init_call)

    # Update verifier parsing: customer_service uses HR-style verifiers
    # The parseHrV function already handles passed=[...]/missing=[...] format
    # We just need to update the conditional in renderRollouts
    old_verifier_check = "if(currentEnv==='hr'){"
    new_verifier_check = "if(currentEnv==='hr'||currentEnv==='customer_service'||currentEnv==='finance'){"
    new_html = new_html.replace(old_verifier_check, new_verifier_check, 1)

    # Fix 1: CS pScore should use parseHrV (x/y format), not parseFinV
    # The metrics row computes pScore — update to include customer_service
    old_pscore = "const pScore=currentEnv==='hr'?(()=>{const c=parseHrV(model.verifier_output);return`${c.passed.length}/${c.passed.length+c.failed.length}`;})():parseFinV(model.verifier_output).scoreStr;"
    new_pscore = "const pScore=(currentEnv==='hr'||currentEnv==='customer_service'||currentEnv==='finance')?(()=>{const c=parseHrV(model.verifier_output);return`${c.passed.length}/${c.passed.length+c.failed.length}`;})():parseFinV(model.verifier_output).scoreStr;"
    new_html = new_html.replace(old_pscore, new_pscore)

    old_ppct = r"const pPct=currentEnv==='hr'?(()=>{const c=parseHrV(model.verifier_output);return c.passed.length/(c.passed.length+c.failed.length);})():(()=>{const m=pScore.match(/(\d+)\/(\d+)/);return m?+m[1]/+m[2]:0;})();"
    new_ppct = "const pPct=(currentEnv==='hr'||currentEnv==='customer_service'||currentEnv==='finance')?(()=>{const c=parseHrV(model.verifier_output);return c.passed.length/(c.passed.length+c.failed.length);})():(()=>{const m=pScore.match(/(\\d+)\\/(\\d+)/);return m?+m[1]/+m[2]:0;})();"
    new_html = new_html.replace(old_ppct, new_ppct)

    # Fix 2: CS environment view should show profiles (like HR)
    old_render_env = """function renderEnvironment(){
  const env=DATA[currentEnv],bar=document.getElementById('env-tabs-bar');
  if(currentEnv==='hr'){
    bar.innerHTML=['Employee Profiles','Org Chart','Policies & Docs'].map((l,i)=>
      `<button class="env-tab-btn${i===0?' active':''}" onclick="switchET(this,'${['profiles','orgchart','policies'][i]}')">${l}</button>`
    ).join('');
    renderET('profiles');
  }else{
    bar.innerHTML=['Golden Reference','Template'].map((l,i)=>
      `<button class="env-tab-btn${i===0?' active':''}" onclick="switchET(this,'${['golden','template'][i]}')">${l}</button>`
    ).join('');
    renderET('golden');
  }
}"""
    new_render_env = """function renderEnvironment(){
  const env=DATA[currentEnv],bar=document.getElementById('env-tabs-bar');
  if(currentEnv==='hr'||currentEnv==='customer_service'){
    var label=currentEnv==='hr'?'Employee Profiles':'Profiles';
    bar.innerHTML=[label,'Policies & Docs'].map((l,i)=>
      `<button class="env-tab-btn${i===0?' active':''}" onclick="switchET(this,'${['profiles','policies'][i]}')">${l}</button>`
    ).join('');
    renderET('profiles');
  }else if(currentEnv==='finance'){
    bar.innerHTML=`<button class="env-tab-btn active" onclick="switchET(this,'policies')">Documents</button>`;
    renderET('policies');
  }else{
    bar.innerHTML=['Golden Reference','Template'].map((l,i)=>
      `<button class="env-tab-btn${i===0?' active':''}" onclick="switchET(this,'${['golden','template'][i]}')">${l}</button>`
    ).join('');
    renderET('golden');
  }
}"""
    new_html = new_html.replace(old_render_env, new_render_env)

    # Also update renderET to handle CS profiles
    old_render_et_check = "if(currentEnv==='hr'){"
    # This appears twice — in renderEnvironment (already fixed) and renderET
    # We need to update the one in renderET (line ~838)
    new_html = new_html.replace(
        "function renderET(tabId){\n  const env=DATA[currentEnv],c=document.getElementById('env-content');let h='';\n  if(currentEnv==='hr'){",
        "function renderET(tabId){\n  const env=DATA[currentEnv],c=document.getElementById('env-content');let h='';\n  if(currentEnv==='hr'||currentEnv==='customer_service'||currentEnv==='finance'){",
    )

    # Update back-link and sidebar title
    new_html = new_html.replace('Collinear Showcase', 'Simulation Lab Showcase')

    # Move theme toggle from back-link-bar to sidebar bottom
    new_html = new_html.replace(
        '<button class="theme-toggle-btn" id="theme-toggle" title="Toggle dark mode">&#9790;</button>',
        ''
    )
    # Add night mode toggle at bottom of sidebar
    night_mode_html = '''<div class="sidebar-night-mode" id="night-mode-area">
    <span>Night Mode</span>
    <label class="toggle-switch"><input type="checkbox" id="theme-toggle"><span class="toggle-slider"></span></label>
  </div>'''
    new_html = new_html.replace('</aside>', night_mode_html + '\n</aside>')

    # Add night mode CSS
    night_css = """
.sidebar-night-mode { margin-top:auto; padding:16px; border-top:1px solid var(--g200); display:flex; align-items:center; justify-content:space-between; font-size:12px; font-weight:500; color:var(--g500); }
html.dark .sidebar-night-mode { border-color:#333; color:#888; }
.toggle-switch { position:relative; width:36px; height:20px; flex-shrink:0; }
.toggle-switch input { opacity:0; width:0; height:0; }
.toggle-slider { position:absolute; top:0; left:0; right:0; bottom:0; background:var(--g300); border-radius:20px; cursor:pointer; transition:.2s; }
.toggle-slider:before { content:''; position:absolute; height:14px; width:14px; left:3px; bottom:3px; background:white; border-radius:50%; transition:.2s; }
.toggle-switch input:checked+.toggle-slider { background:var(--orange); }
.toggle-switch input:checked+.toggle-slider:before { transform:translateX(16px); }
html.dark .toggle-slider { background:#555; }
"""
    new_html = new_html.replace('</style>', night_css + '</style>', 1)

    # Update theme toggle JS to use checkbox instead of button
    new_html = new_html.replace(
        "btn.addEventListener('click', function(){",
        "btn.addEventListener('change', function(){"
    )
    new_html = new_html.replace(
        "btn.innerHTML = document.documentElement.classList.contains('dark') ? '&#9728;' : '&#9790;';",
        "btn.checked = document.documentElement.classList.contains('dark');"
    )
    new_html = new_html.replace(
        "this.innerHTML = isDark ? '&#9728;' : '&#9790;';",
        "this.checked = isDark;"
    )
    new_html = new_html.replace(
        '<div class="brand-text">Collinear AI<span>Enterprise</span></div>',
        '<div class="brand-text">Collinear AI<span>Enterprise Workflows</span></div>'
    )
    new_html = new_html.replace(
        '<title>Collinear AI — Enterprise</title>',
        '<title>Collinear AI — Enterprise Workflows</title>'
    )

    # Remove cost metric from rollouts
    new_html = new_html.replace(
        """<div class="metric-box"><div class="metric-label">Cost</div><div class="metric-val">$${meta.cost_usd}</div></div>""",
        ""
    )

    # Move model tabs from main header to sidebar
    # 1. Remove from main header
    new_html = new_html.replace(
        '<div class="main-header"><div class="model-tabs" id="mtabs"></div></div>',
        '<div class="main-header" id="main-header-bar" style="display:none"></div>'
    )

    # 2. Add model tabs container in sidebar, after task pills
    model_sidebar_html = '''  <div class="sidebar-divider" id="model-divider"></div>
  <div class="sidebar-section">Model</div>
  <div class="model-pills" id="mtabs" style="padding:4px 16px 8px"></div>'''
    new_html = new_html.replace(
        '  <div class="sidebar-divider"></div>\n  <nav class="sidebar-nav">',
        model_sidebar_html + '\n  <div class="sidebar-divider"></div>\n  <nav class="sidebar-nav">'
    )

    # 3. Override model-tabs CSS to be vertical in sidebar
    model_sidebar_css = """
.model-pills { display:flex; flex-direction:column; gap:6px; }
#mtabs .model-tab { display:block; width:100%; text-align:left; padding:8px 12px; border-radius:8px; font-size:12.5px; }
"""
    new_html = new_html.replace(
        '</style>',
        model_sidebar_css + '</style>',
        1
    )

    # Define curTask at the top of renderRollouts so brief/output_format are available
    new_html = new_html.replace(
        "function renderRollouts(){\n  const env=DATA[currentEnv],model=env.models[currentModel],meta=model.meta,rubric=parseRubric(model.rubric_output);\n  let h='';",
        "function renderRollouts(){\n  const env=DATA[currentEnv],model=env.models[currentModel],meta=model.meta,rubric=parseRubric(model.rubric_output);\n  var curTask=null;if(env.tasks&&env.current_task)curTask=env.tasks[env.current_task];\n  let h='';",
    )

    # Fix 3: Rename "LLM ANNOTATION" to "What Happened?" and collapse by default
    new_html = new_html.replace(
        "LLM ANNOTATION &mdash; ${esc(MN[currentModel]||currentModel)}",
        "What Happened?"
    )
    # Collapse annotation by default (was max-height:5000px)
    new_html = new_html.replace(
        """<div class="collapsible" id="annot" style="max-height:5000px">""",
        """<div class="collapsible" id="annot" style="max-height:0">"""
    )

    # Fix 3b: Collapse rubric scores by default (was max-height:5000px)
    new_html = new_html.replace(
        """<div class="collapsible" id="rdims" style="max-height:5000px">""",
        """<div class="collapsible" id="rdims" style="max-height:0">"""
    )

    # Fix 4: Render task as markdown, make it collapsible, with brief visible in header
    new_html = new_html.replace(
        """<div class="card"><div class="card-title"><div class="dot" style="background:var(--orange)"></div>TASK</div><div class="prompt-box">${esc(userTask)}</div></div>""",
        """<div class="rsection"><div class="rsection-header" onclick="toggle('taskdesc')" style="flex-wrap:wrap"><h3>TASK DESCRIPTION</h3><span class="chevron-toggle" id="taskdesc-ch">&#9660;</span>${(function(){var b='';if(curTask&&curTask.brief)b='<div style="width:100%;font-size:12.5px;font-weight:400;color:var(--g500);margin-top:4px;line-height:1.5">'+esc(curTask.brief)+'</div>';return b;})()}</div><div class="collapsible" id="taskdesc" style="max-height:0"><div style="padding:16px 20px">${md(userTask)}</div></div></div>${(function(){var ts=new Set();model.steps.forEach(function(s){if(s.type==='tool_call'&&s.tool){var n=s.tool.indexOf('__')>=0?s.tool.split('__').pop():s.tool;n=n.replace(/_/g,' ');ts.add(n);}});if(!ts.size)return '';var h='<div style="display:flex;flex-wrap:wrap;gap:6px;padding:8px 0;align-items:center"><span style="font-size:11px;font-weight:600;color:var(--g400);text-transform:uppercase;letter-spacing:.5px;margin-right:4px">Tools</span>';ts.forEach(function(t){h+='<span style="padding:3px 10px;border-radius:12px;font-size:11px;background:var(--g100);color:var(--g600);white-space:nowrap">'+t+'</span>';});h+='</div>';return h;})()}"""
    )

    # Fix 5: Add output format collapsible and output xlsx tables for finance
    # Insert after the TRAJECTORY section in renderRollouts
    finance_output_js = r"""
// Output requirements collapsible (finance)
if(curTask&&curTask.output_format){
  h+='<div class="rsection"><div class="rsection-header" onclick="toggle(\'outfmt\')"><h3>EXPECTED OUTPUT</h3><span class="chevron-toggle" id="outfmt-ch">&#9660;</span></div><div class="collapsible" id="outfmt" style="max-height:0"><div style="padding:16px 20px">'+md(curTask.output_format)+'</div></div></div>';
}
// Agent output xlsx tables (finance)
if(model.output_xlsx&&Object.keys(model.output_xlsx).length>0){
  var oxPfx='oxs-'+currentModel+'-';
  h+='<div class="rsection"><div class="rsection-header" onclick="toggle(\'outxlsx\')"><h3>AGENT OUTPUT</h3><span class="badge badge-lg" style="background:#e8f5e9;color:#2e7d32">'+Object.keys(model.output_xlsx).length+' sheets</span><span class="chevron-toggle" id="outxlsx-ch">&#9660;</span></div><div class="collapsible" id="outxlsx" style="max-height:0"><div style="padding:16px 20px">';
  var sheetKeys=Object.keys(model.output_xlsx);
  h+='<div style="display:flex;gap:6px;margin-bottom:12px">'+sheetKeys.map(function(sn,i){return '<button class="model-tab'+(i===0?' active':'')+'" onclick="showXlsxSheet(this,\''+oxPfx+i+'\')">'+esc(sn)+'</button>';}).join('')+'</div>';
  sheetKeys.forEach(function(sn,i){
    var rows=model.output_xlsx[sn];
    h+='<div id="'+oxPfx+i+'" style="'+(i>0?'display:none;':'')+'max-height:400px;overflow:auto"><table class="seed-table"><tbody>';
    rows.forEach(function(row,ri){
      var tag=ri===0?'th':'td';
      h+='<tr>'+row.map(function(c){return '<'+tag+(typeof c==='number'?' class="num"':'')+'>'+esc(String(c))+'</'+tag+'>';}).join('')+'</tr>';
    });
    h+='</tbody></table></div>';
  });
  h+='</div></div></div>';
}
"""
    new_html = new_html.replace(
        "document.getElementById('rollouts-content').innerHTML=h;",
        finance_output_js + "document.getElementById('rollouts-content').innerHTML=h;"
    )

    # Fix 6: Add xlsx sheet tab switching helper and seed-table CSS
    xlsx_helpers = r"""
function showXlsxSheet(btn,id){
  btn.parentElement.querySelectorAll('.model-tab').forEach(function(b){b.classList.remove('active');});
  btn.classList.add('active');
  var parent=btn.parentElement.parentElement;
  var pfx=id.replace(/\d+$/,'');
  parent.querySelectorAll('[id^="'+pfx+'"]').forEach(function(el){el.style.display='none';});
  document.getElementById(id).style.display='';
}
"""
    new_html = new_html.replace(
        'function switchEnv(env){',
        xlsx_helpers + 'function switchEnv(env){'
    )

    # Fix 7: Update finance environment view to show input xlsx tables
    new_render_env_fin = """  }else if(currentEnv==='finance'){
    bar.innerHTML=`<button class="env-tab-btn active" onclick="switchET(this,'fin_inputs')">Input Files</button>`;
    renderET('fin_inputs');
  }else{"""
    new_html = new_html.replace(
        """  }else if(currentEnv==='finance'){
    bar.innerHTML=`<button class="env-tab-btn active" onclick="switchET(this,'policies')">Documents</button>`;
    renderET('policies');
  }else{""",
        new_render_env_fin
    )

    # Add finance input xlsx rendering in renderET
    fin_et_js = r"""
  if(tabId==='fin_inputs'){
    var curTask=null;
    if(env.tasks&&env.current_task)curTask=env.tasks[env.current_task];
    if(curTask&&curTask.input_xlsx){
      var files=curTask.input_xlsx;
      Object.keys(files).forEach(function(fname){
        var sheets=files[fname];
        h+='<div style="margin-bottom:16px"><h4 style="font-size:13px;font-weight:600;margin-bottom:8px;color:var(--g700)">'+esc(fname)+'</h4>';
        var snames=Object.keys(sheets);
        if(snames.length>1){
          h+='<div style="display:flex;gap:6px;margin-bottom:8px">'+snames.map(function(sn,i){return '<button class="model-tab'+(i===0?' active':'')+'" onclick="showFinSheet(this,\'fis-'+fname.replace(/\W/g,'')+'-'+i+'\')">'+esc(sn)+'</button>';}).join('')+'</div>';
        }
        snames.forEach(function(sn,i){
          var rows=sheets[sn];
          h+='<div id="fis-'+fname.replace(/\W/g,'')+'-'+i+'" style="'+(i>0?'display:none;':'')+'max-height:400px;overflow:auto"><table class="seed-table"><tbody>';
          rows.forEach(function(row,ri){
            var tag=ri===0?'th':'td';
            h+='<tr>'+row.map(function(c){return '<'+tag+(typeof c==='number'?' class="num"':'')+'>'+esc(String(c))+'</'+tag+'>';}).join('')+'</tr>';
          });
          h+='</tbody></table></div>';
        });
        h+='</div>';
      });
    }else{
      h+='<p style="color:var(--g500);padding:20px">No input files available for this task.</p>';
    }
    c.innerHTML=h;return;
  }
"""
    new_html = new_html.replace(
        "function renderET(tabId){\n  const env=DATA[currentEnv],c=document.getElementById('env-content');let h='';\n  if(currentEnv==='hr'||currentEnv==='customer_service'||currentEnv==='finance'){",
        "function renderET(tabId){\n  const env=DATA[currentEnv],c=document.getElementById('env-content');let h='';\n" + fin_et_js + "  if(currentEnv==='hr'||currentEnv==='customer_service'){",
    )

    # Fix 9: Add showFinSheet helper
    fin_sheet_helper = r"""
function showFinSheet(btn,id){
  btn.parentElement.querySelectorAll('.model-tab').forEach(function(b){b.classList.remove('active');});
  btn.classList.add('active');
  var parent=btn.closest('div[style*="margin-bottom"]')||btn.parentElement.parentElement;
  parent.querySelectorAll('[id^="fis-"]').forEach(function(el){el.style.display='none';});
  document.getElementById(id).style.display='';
}
"""
    new_html = new_html.replace(
        'function showXlsxSheet',
        fin_sheet_helper + 'function showXlsxSheet'
    )

    # Fix 10: Also propagate output_format, brief, input_xlsx in switchTask
    new_html = new_html.replace(
        "env.task=t.task;env.rubric_definition=t.rubric_definition;\n  env.models=t.models;env.npc_profiles=t.npc_profiles||env.npc_profiles;",
        "env.task=t.task;env.rubric_definition=t.rubric_definition;env.output_format=t.output_format||'';env.brief=t.brief||'';\n  env.models=t.models;env.npc_profiles=t.npc_profiles||env.npc_profiles;",
        1  # only first occurrence (in switchTask)
    )

    # Write output
    with open(INDEX, 'w') as f:
        f.write(new_html)

    size_kb = len(new_html.encode('utf-8')) / 1024
    print(f"\nWritten {INDEX} ({size_kb:.0f} KB)")


if __name__ == '__main__':
    main()
